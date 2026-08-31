"""Export the outreach workbook -- who to contact, by which route.

Stage 4's workbooks answer "may I email this shop". This one answers "do I
have a way in, and is this shop even the right shape for the pitch", so it
cuts across tiers: a German shop that may never be cold-emailed is still a
perfectly good LinkedIn approach once the Impressum names its owner.

Because it crosses tiers, every row carries an explicit `email_ok` column.
Never send to a row marked no -- the tier rules still hold.

    python -m pipeline.export_outreach [--min-images 20]
"""
from __future__ import annotations

import argparse
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config, db
from .cli import heading, table
from .stage4_score import clean_cell

COLUMNS = [
    ("domain", 30), ("email", 32), ("email_ok", 9), ("contact_route", 14),
    ("owner_name", 24), ("linkedin_url", 44), ("linkedin_kind", 13),
    ("tier", 14), ("score", 7), ("platform", 13), ("jurisdiction", 12),
    ("created_year", 12), ("image_count", 11), ("images_no_alt", 13),
    ("sample_image_kb", 15), ("page_title", 44), ("final_url", 40),
]

SELECT = """
SELECT domain, email,
       CASE WHEN tier = 'tier1_email' THEN 'yes' ELSE 'NO' END AS email_ok,
       CASE WHEN owner_name IS NOT NULL AND linkedin_url IS NOT NULL THEN 'name+linkedin'
            WHEN owner_name IS NOT NULL THEN 'named owner'
            WHEN linkedin_kind = 'personal' THEN 'linkedin person'
            WHEN linkedin_url IS NOT NULL THEN 'linkedin company'
            ELSE 'email only' END AS contact_route,
       owner_name, linkedin_url, linkedin_kind, tier, score, platform,
       jurisdiction, created_year, image_count, images_no_alt,
       CASE WHEN sample_image_bytes IS NULL THEN NULL
            ELSE sample_image_bytes / 1024 END AS sample_image_kb,
       page_title, final_url
  FROM targets
"""

# Ordered best-first within each sheet: a shop with a named human and a lot of
# images is worth more than a high score with no way in.
ORDER = (" ORDER BY (owner_name IS NOT NULL) DESC, "
         "(linkedin_kind = 'personal') DESC, image_count DESC, score DESC")


def sheets(min_images: int) -> list[tuple[str, str, str]]:
    """(sheet name, note, WHERE clause) -- most actionable first."""
    return [
        ("best leads",
         f"Image-heavy shops with a working cart AND a way in: {min_images}+ "
         "images, plus a name or LinkedIn. Work this sheet first.",
         f"image_count >= {min_images} AND has_cart = 1 AND tier != 'dropped' "
         "AND (owner_name IS NOT NULL OR linkedin_url IS NOT NULL)"),
        ("image sites no cart",
         f"{min_images}+ images and a contact, but no cart -- photographers, "
         "galleries, catalogues and portfolios. Not shops, so they never "
         "qualified for a tier, but they own a lot of images and may be the "
         "strongest watermarking prospects on the list. Judge separately.",
         f"image_count >= {min_images} AND (has_cart = 0 OR has_cart IS NULL) "
         "AND (owner_name IS NOT NULL OR linkedin_url IS NOT NULL)"),
        ("named owners",
         "A real person's name, mostly from the German/Austrian/Swiss Impressum. "
         "Many are EU shops you may NOT email -- LinkedIn or post only.",
         "owner_name IS NOT NULL"),
        ("linkedin only",
         "A LinkedIn account but no name. Company pages need one more hop to "
         "reach a human.",
         "linkedin_url IS NOT NULL AND owner_name IS NULL"),
        ("image heavy no contact",
         f"{min_images}+ images but no name and no LinkedIn. Good fit, email is "
         "the only route -- and only where email_ok says yes.",
         f"image_count >= {min_images} AND owner_name IS NULL AND linkedin_url IS NULL"),
    ]


def write_sheet(wb, conn, name, note, where) -> int:
    ws = wb.create_sheet(name[:31])
    warn = openpyxl.cell.WriteOnlyCell(ws, value=note)
    warn.font = Font(bold=True, italic=True)
    ws.append([warn])
    ws.append([openpyxl.cell.WriteOnlyCell(ws, value="")])

    header = []
    for col, _ in COLUMNS:
        cell = openpyxl.cell.WriteOnlyCell(ws, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
        header.append(cell)
    ws.append(header)

    red = PatternFill("solid", fgColor="FFF0F0")
    rows = 0
    for row in conn.execute(SELECT + " WHERE profile_checked_at IS NOT NULL AND "
                            + where + ORDER):
        out = []
        for col, _ in COLUMNS:
            cell = openpyxl.cell.WriteOnlyCell(ws, value=clean_cell(row[col]))
            # Make "do not email this one" impossible to miss at a glance.
            if col == "email_ok" and row["email_ok"] == "NO":
                cell.font = Font(bold=True, color="B00020")
                cell.fill = red
            out.append(cell)
        ws.append(out)
        rows += 1

    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A4"
    return rows


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description=__doc__.strip().splitlines()[0])
    p.add_argument("--db", default=config.DB_PATH)
    p.add_argument("--out-dir", default=config.OUT_DIR)
    p.add_argument("--min-images", type=int, default=20,
                   help="images on the home page before a shop counts as "
                        "image-heavy retail (default: %(default)s)")
    args = p.parse_args(argv)

    conn = db.connect(args.db)
    os.makedirs(args.out_dir, exist_ok=True)
    path = os.path.join(args.out_dir, "outreach.xlsx")

    wb = openpyxl.Workbook(write_only=True)
    heading("Outreach workbook")
    counts = []
    for name, note, where in sheets(args.min_images):
        counts.append((name, write_sheet(wb, conn, name, note, where)))
    wb.save(path)

    print(table([(n, f"{c:,}") for n, c in counts], ["sheet", "rows"]))
    print(f"\nwritten to {path}")
    print("\nEvery sheet carries email_ok. A row marked NO is tier 2 or 3 -- "
          "reachable on LinkedIn, never by cold email.")
    conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
