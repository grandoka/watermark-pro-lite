"""Stage 4 -- score every target 0-100 and split it into outreach tiers.

The score answers one question: how likely is this to be a *reachable,
fixable, recent* store? Platform fit dominates, because that is what decides
whether the product can help at all; commerce signals and recency come next;
a shared role inbox is a small penalty, since those reply less often.

Tiering is a legal split before it is a quality split. Only tier 1 is ever
cold-emailed. EU and Canadian domains, and freemail contacts, go to tier 2 --
their addresses are for paid-audience upload (Meta / Google Customer Match,
lookalike seeding), never for sending.

    python -m pipeline.stage4_score [--limit N] [--min-score 55]
"""
from __future__ import annotations

import datetime as dt
import os

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from . import config, db
from .cli import base_parser, heading, table

TIER1, TIER2, TIER3, DROPPED = "tier1_email", "tier2_ads", "tier3_nurture", "dropped"

# Columns every output workbook carries, in order.
OUTPUT_COLUMNS = ["domain", "email", "platform", "score", "created_year",
                  "jurisdiction", "http_status", "has_cart", "page_title",
                  "final_url"]

WORKBOOKS = {
    TIER1: ("tier1_email.xlsx", "Cold email -- the only tier cleared to send"),
    TIER2: ("tier2_ads.xlsx", "Paid audience upload ONLY -- do not send email"),
    TIER3: ("tier3_nurture.xlsx", "Hold for later -- weak fit or not crawled"),
    DROPPED: ("dropped.xlsx", "Excluded, with the reason"),
}


def parse_args(argv=None):
    p = base_parser(__doc__.strip().splitlines()[0])
    p.add_argument("--min-score", type=int, default=55, metavar="N",
                   help="score at or above which a live domain is worth "
                        "contacting now rather than nurturing (default: %(default)s)")
    p.add_argument("--out-dir", default=config.OUT_DIR,
                   help="directory for the output workbooks (default: %(default)s)")
    p.add_argument("--no-workbooks", action="store_true",
                   help="score and tier in the database, skip the .xlsx export")
    return p.parse_args(argv)


# --- scoring ---------------------------------------------------------------

def recency_points(year: int | None) -> int:
    if year is None:
        return 0
    if year >= 2024:
        return 15
    if year >= 2022:
        return 12
    if year >= 2020:
        return 6
    return 0


def score_row(row) -> int:
    """0-100. Only ever called for rows that reached a live 200 response."""
    score = 5  # reached a 200; everything below builds on that

    # Platform fit is the single biggest lever: Shopify and WooCommerce stores
    # are the ones the product actually drops into.
    score += config.PLATFORM_TIERS.get(row["platform"], 0)

    # Commerce liveness. Product schema is rare on a home page (a few percent
    # of stores put it there), so a cart on its own still earns most of this.
    if row["has_cart"]:
        score += 12
    if row["has_product_schema"]:
        score += 13

    score += recency_points(row["created_year"])

    if row["has_mx"]:
        score += 15  # a mailbox probably exists behind the address

    if row["response_time_ms"] is not None and row["response_time_ms"] < 2000:
        score += 5  # a site that answers fast is a site someone maintains

    if row["is_role"]:
        score -= 5  # shared inbox, lower reply rate

    return max(0, min(100, score))


def disqualify(row) -> str | None:
    """Reason this target can never be emailed, or None."""
    if row["status"] == config.STATUS_FREEMAIL:
        return None  # not disqualified, just ads-only -- handled in tiering
    if row["status"] == config.STATUS_DEAD:
        return "does not resolve or nothing served"
    if row["status"] == config.STATUS_MX_ONLY:
        return "mail exchanger but no website"
    if row["status"] == config.STATUS_BLOCKED:
        return "robots.txt disallows the site root"
    if row["status"] == config.STATUS_PARKED or row["is_parked"]:
        return "parked or for-sale page"
    if row["status"] == config.STATUS_AGED_OUT:
        return None  # deliberately not crawled -- nurture, not dropped
    if row["status"] != config.STATUS_LIVE:
        return f"not live (status {row['status']})"
    if row["http_status"] != 200:
        return f"non-200 response ({row['http_status']})"
    if not row["has_cart"]:
        return "no cart or checkout link found"
    return None


def assign(row, min_score: int) -> tuple[str | None, int | None, str | None]:
    """Return (tier, score, reason) for one target.

    A target the earlier stages have not reached yet gets no tier at all --
    it is unknown, not rejected, and putting it in dropped.xlsx would quietly
    turn "we have not looked" into "we looked and it was bad".
    """
    if row["status"] == config.STATUS_PENDING:
        return None, None, None

    if row["status"] == config.STATUS_FREEMAIL:
        # No store site to judge, so no score -- these exist purely as seed
        # addresses for lookalike audiences.
        return TIER2, 0, "freemail contact: ads audience only, never email"

    if row["status"] == config.STATUS_AGED_OUT:
        return TIER3, 0, "contact predates the crawl cutoff; not fetched"

    reason = disqualify(row)
    if reason:
        return DROPPED, 0, reason

    score = score_row(row)
    if score < min_score:
        return TIER3, score, f"live but scores {score} (below {min_score})"

    if row["is_freemail"]:
        return TIER2, score, "freemail contact: ads audience only, never email"
    if row["jurisdiction"] not in config.EMAILABLE_JURISDICTIONS:
        return TIER2, score, (f"{row['jurisdiction']} jurisdiction: "
                              "ads audience only, no cold email")
    if not row["has_mx"]:
        return TIER2, score, "no MX record: cannot be emailed, ads audience only"
    return TIER1, score, "cleared for cold email"


UPDATE = ("UPDATE targets SET score = :score, tier = :tier, tier_reason = :reason, "
          "scored_at = :scored_at WHERE target_key = :key")


SCORE_CHUNK = 5000


def score_all(conn, args) -> int:
    """Score and tier every target, in chunks.

    Reads through a second connection: writing to a table while iterating a
    cursor over it on the same connection is not safe, and the full four-file
    set is ~600k rows, which is more than is worth holding in memory at once.
    """
    reader = db.connect(args.db)
    sql = "SELECT * FROM targets"
    if args.limit:
        sql += f" LIMIT {int(args.limit)}"
    now = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")

    cursor = reader.execute(sql)
    scored = 0
    while True:
        rows = cursor.fetchmany(SCORE_CHUNK)
        if not rows:
            break
        updates = []
        for row in rows:
            tier, score, reason = assign(row, args.min_score)
            updates.append({"key": row["target_key"], "tier": tier, "score": score,
                            "reason": reason, "scored_at": now})
        conn.executemany(UPDATE, updates)
        conn.commit()
        scored += len(updates)
    reader.close()
    return scored


# --- workbooks -------------------------------------------------------------

def write_workbook(conn, tier: str, path: str, note: str) -> int:
    columns = list(OUTPUT_COLUMNS)
    if tier == DROPPED:
        columns.append("reason")
    select = ", ".join("tier_reason AS reason" if c == "reason" else c
                       for c in columns)

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(tier)

    bold = Font(bold=True)
    title_cell = openpyxl.cell.WriteOnlyCell(ws, value=note)
    title_cell.font = Font(bold=True, italic=True)
    ws.append([title_cell])
    ws.append([openpyxl.cell.WriteOnlyCell(ws, value="")])

    head = []
    for name in columns:
        cell = openpyxl.cell.WriteOnlyCell(ws, value=name)
        cell.font = bold
        cell.alignment = Alignment(horizontal="left")
        head.append(cell)
    ws.append(head)

    rows = 0
    for row in conn.execute(
            f"SELECT {select} FROM targets WHERE tier = ? "
            "ORDER BY score DESC, created_year DESC, domain", (tier,)):
        ws.append([row[c] for c in columns])
        rows += 1

    widths = {"domain": 32, "email": 34, "platform": 14, "score": 7,
              "created_year": 13, "jurisdiction": 13, "http_status": 12,
              "has_cart": 9, "page_title": 50, "final_url": 46, "reason": 44}
    for i, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 16)
    ws.freeze_panes = "A4"

    wb.save(path)
    return rows


def export(conn, out_dir: str) -> dict[str, int]:
    os.makedirs(out_dir, exist_ok=True)
    written = {}
    for tier, (filename, note) in WORKBOOKS.items():
        path = os.path.join(out_dir, filename)
        written[tier] = write_workbook(conn, tier, path, note)
        print(f"  {filename:<22} {written[tier]:>8,} rows")
    return written


# --- summary ---------------------------------------------------------------

def print_summary(conn, args) -> None:
    heading("Stage 4 summary")

    total = db.count(conn)
    order = [TIER1, TIER2, TIER3, DROPPED]
    counts = dict(db.summarize(conn, "tier"))
    scored_total = sum(counts.get(t, 0) for t in order)
    print(table([(t, f"{counts.get(t, 0):,}",
                  f"{100*counts.get(t, 0)/max(scored_total, 1):.1f}%")
                 for t in order],
                ["tier", "targets", "share"]))
    for t in order:
        print(f"  {t:<14} {WORKBOOKS[t][1]}")
    unprocessed = total - scored_total
    if unprocessed:
        print(f"\n{unprocessed:,} targets are not yet tiered -- the earlier stages "
              f"have not reached them. They are in no workbook.")

    live = db.count(conn, f"status = '{config.STATUS_LIVE}'")
    if live:
        print(f"\nPlatform distribution among {live:,} live domains:")
        rows = db.summarize(conn, "platform", f"status = '{config.STATUS_LIVE}'")
        print(table([(k or "(unidentified)", f"{n:,}", f"{100*n/live:.1f}%")
                     for k, n in rows], ["platform", "domains", "share"]))

    t1 = counts.get(TIER1, 0)
    if t1:
        print(f"\nTop 20 TLDs in tier 1 ({t1:,} domains):")
        rows = conn.execute(
            "SELECT tld, COUNT(*) n FROM targets WHERE tier = ? "
            "GROUP BY 1 ORDER BY n DESC LIMIT 20", (TIER1,)).fetchall()
        print(table([(r["tld"], f"{r['n']:,}", f"{100*r['n']/t1:.1f}%") for r in rows],
                    ["tld", "domains", "share"]))
        stats = conn.execute(
            "SELECT AVG(score) a, MIN(score) mn, MAX(score) mx FROM targets "
            "WHERE tier = ?", (TIER1,)).fetchone()
        print(f"\nTier 1 score: min {stats['mn']}, mean {stats['a']:.1f}, max {stats['mx']}")

    dropped = conn.execute(
        "SELECT tier_reason, COUNT(*) n FROM targets WHERE tier = ? "
        "GROUP BY 1 ORDER BY n DESC LIMIT 10", (DROPPED,)).fetchall()
    if dropped:
        print("\nWhy targets were dropped:")
        print(table([(r["tier_reason"], f"{r['n']:,}") for r in dropped],
                    ["reason", "targets"]))

    uncrawled = db.count(conn, "is_freemail = 0 AND dns_resolves = 1 "
                               "AND http_checked_at IS NULL "
                               f"AND status != '{config.STATUS_AGED_OUT}'")
    if uncrawled:
        print(f"\nNote: {uncrawled:,} resolvable domains have not been fetched yet. "
              f"Run stage 3 to completion and re-run stage 4 for final numbers.")


def main(argv=None) -> int:
    args = parse_args(argv)
    conn = db.connect(args.db)
    heading(f"Stage 4 -- scoring (tier 1 cutoff: {args.min_score})")
    scored = score_all(conn, args)
    print(f"  scored {scored:,} targets")
    if not args.no_workbooks:
        print()
        export(conn, args.out_dir)
    print_summary(conn, args)
    conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
