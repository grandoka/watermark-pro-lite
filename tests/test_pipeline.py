"""Tests for the pure logic: extraction, classification, fingerprinting, tiering.

Nothing here touches the network. The stages that do are thin wrappers around
these functions, so this is where the behaviour that decides who gets emailed
is pinned down.
"""
from __future__ import annotations

import asyncio
import datetime as dt
import sqlite3

import dns.exception
import dns.resolver
import pytest

from pipeline import config, db
from pipeline.stage1_ingest import classify, clean_email, row_created, _is_header
from pipeline.stage2_dns import check_domain
from pipeline.stage3_http import (candidate_urls, detect_platform, extract_title,
                                  fingerprint)
from pipeline.stage4_score import (DROPPED, TIER1, TIER2, TIER3, assign,
                                   recency_points, score_row)


# --- stage 1: extraction ---------------------------------------------------

@pytest.mark.parametrize("raw, expected", [
    ("  Info@Example.COM ", "info@example.com"),
    ("mailto:sales@shop.de", "sales@shop.de"),
    ("contact@store.co.uk,", "contact@store.co.uk"),
    ("<hola@tienda.es>", "hola@tienda.es"),
    ("not an email", None),
    ("broken@nodot", None),
    ("two@@at.com", None),
    ("spaced @ out.com", None),
])
def test_clean_email(raw, expected):
    assert clean_email(raw) == expected


def test_scanner_splits_packed_cells():
    """Source cells pack several addresses together with colons."""
    cell = "corporatestationbd@gmail.com:info@corporatestationbd.com"
    found = [clean_email(m) for m in config.EMAIL_SCAN.findall(cell)]
    assert found == ["corporatestationbd@gmail.com", "info@corporatestationbd.com"]


def test_header_row_detected_only_when_it_is_one():
    assert _is_header(("created", "emails"))
    # The headerless sheets start straight in on data.
    assert not _is_header(("info@hearfit.ca", "support@hearfit.ca"))
    assert not _is_header((None, None))


def test_row_created_finds_the_date_in_any_position():
    stamp = dt.datetime(2024, 6, 7)
    assert row_created(("a@b.com", stamp)) == stamp
    assert row_created((stamp, "a@b.com")) == stamp
    assert row_created(("a@b.com", "c@d.com")) is None


# --- stage 1: classification -----------------------------------------------

def test_store_domain_is_keyed_by_domain():
    rec = classify("info@mystore.com")
    assert rec["target_key"] == "mystore.com"
    assert rec["is_freemail"] == 0
    assert rec["is_role"] == 1
    assert rec["status"] == config.STATUS_PENDING


def test_freemail_is_keyed_by_address_so_contacts_are_not_collapsed():
    """26k gmail contacts must stay 26k rows, not become one gmail.com row."""
    a = classify("shop.one@gmail.com")
    b = classify("shop.two@gmail.com")
    assert a["target_key"] != b["target_key"]
    assert a["domain"] == b["domain"] == "gmail.com"
    assert a["is_freemail"] == 1
    assert a["status"] == config.STATUS_FREEMAIL


@pytest.mark.parametrize("domain, freemail", [
    ("gmail.com", True), ("web.de", True), ("mail.ru", True),
    ("gmx.de", True), ("gmx.net", True), ("yandex.ru", True),
    ("t-online.de", True), ("proton.me", True), ("libero.it", True),
    ("mystore.com", False), ("gmail-store.com", False), ("mailbox-shop.de", False),
])
def test_freemail_detection(domain, freemail):
    assert config.is_freemail_domain(domain) is freemail


@pytest.mark.parametrize("email, jurisdiction", [
    ("a@shop.de", "EU"), ("a@shop.fr", "EU"), ("a@shop.eu", "EU"),
    ("a@shop.ca", "CASL"),
    ("a@shop.uk", "PECR"), ("a@shop.co.uk", "PECR"),
    ("a@shop.com", "PERMISSIVE"), ("a@shop.com.au", "PERMISSIVE"),
    ("a@shop.store", "PERMISSIVE"),
    ("a@shop.ru", "OTHER"), ("a@shop.br", "OTHER"),
])
def test_jurisdiction(email, jurisdiction):
    assert classify(email)["jurisdiction"] == jurisdiction


# --- stage 1: dedupe by domain, keeping the newest contact -----------------

def upsert(conn, email, created):
    from pipeline.stage1_ingest import UPSERT
    rec = classify(email)
    rec.update(created=created, created_year=int(created[:4]) if created else None,
               source_file="t.xlsx", source_sheet="Sheet1")
    conn.execute(UPSERT, rec)
    conn.commit()


@pytest.fixture()
def conn(tmp_path):
    connection = db.connect(str(tmp_path / "t.db"))
    yield connection
    connection.close()


def test_domain_dedupe_keeps_the_most_recent_contact(conn):
    upsert(conn, "old@store.com", "2019-01-01 00:00:00")
    upsert(conn, "new@store.com", "2024-05-05 00:00:00")
    upsert(conn, "older@store.com", "2017-01-01 00:00:00")
    row = conn.execute("SELECT * FROM targets").fetchone()
    assert conn.execute("SELECT COUNT(*) FROM targets").fetchone()[0] == 1
    assert row["email"] == "new@store.com"
    assert row["created_year"] == 2024
    assert row["email_count"] == 3


def test_a_dated_contact_beats_an_undated_one(conn):
    upsert(conn, "nodate@store.com", None)
    upsert(conn, "dated@store.com", "2023-03-03 00:00:00")
    assert conn.execute("SELECT email FROM targets").fetchone()[0] == "dated@store.com"


def test_reingest_does_not_clobber_enrichment(conn):
    """Stage 1 must be safe to re-run after hours of crawling."""
    upsert(conn, "info@store.com", "2024-01-01 00:00:00")
    conn.execute("UPDATE targets SET dns_resolves = 1, platform = 'Shopify', "
                 "status = 'live', http_checked_at = '2026-01-01'")
    conn.commit()
    upsert(conn, "info@store.com", "2024-01-01 00:00:00")
    row = conn.execute("SELECT * FROM targets").fetchone()
    assert row["platform"] == "Shopify"
    assert row["status"] == "live"
    assert row["http_checked_at"] == "2026-01-01"


# --- stage 3: fingerprinting ----------------------------------------------

@pytest.mark.parametrize("html, platform", [
    ('<script src="https://cdn.shopify.com/s/x.js">', "Shopify"),
    ('<link href="/wp-content/plugins/woocommerce/a.css">', "WooCommerce"),
    ('<img src="https://cdn11.bigcommerce.com/s-abc/x.png">', "BigCommerce"),
    ('<script>require(["Magento_Ui/js/core"])</script>', "Magento"),
    ('<div id="prestashop" data-x>', "PrestaShop"),
    ('<style>._wixCssVars{}</style>', "Wix"),
    ('<img src="https://static1.squarespace.com/x.png">', "Squarespace"),
    ('<a href="index.php?route=common/home">', "OpenCart"),
    ('<div class="plain old site"></div>', None),
])
def test_platform_detection(html, platform):
    assert detect_platform(html, {}) == platform


def test_image_urls_are_not_magento():
    """A bare 'mage/' needle matches every 'image/' path -- it mislabelled 18%
    of live stores before the signatures became anchored regexes."""
    html = '<img src="/media/image/hero.png"><link href="/assets/images/a.css">'
    assert detect_platform(html, {}) is None


def test_platform_detection_from_headers():
    assert detect_platform("<html></html>", {"x-shopid": "12345"}) == "Shopify"
    assert detect_platform("<html></html>", {"x-wix-request-id": "abc"}) == "Wix"


def test_extract_title_unescapes_and_collapses():
    assert extract_title("<title>\n  Caf&eacute;  &amp;  Shop\n</title>") == "Café & Shop"
    assert extract_title("<html>no title</html>") is None


@pytest.mark.parametrize("html", [
    '<a href="/cart">Cart</a>',
    '<form action="https://shop.com/checkout" method="post">',
    '<a href="/panier">Panier</a>',
    '<button class="add-to-cart">Add</button>',
])
def test_cart_detection(html):
    assert fingerprint(html.encode(), {})["has_cart"] == 1


def test_product_schema_detection():
    body = b'<script type="application/ld+json">{"@type": "Product"}</script>'
    assert fingerprint(body, {})["has_product_schema"] == 1
    assert fingerprint(b"<html></html>", {})["has_product_schema"] == 0


@pytest.mark.parametrize("body", [
    b"<title>This domain is for sale</title>",
    b"<html><body>Buy this domain</body></html>",
    b"<title>Welcome to nginx!</title>",
])
def test_parked_detection(body):
    assert fingerprint(body, {})["is_parked"] == 1


def test_lang_extraction():
    assert fingerprint(b'<html lang="de-DE">', {})["lang"] == "de-de"


def test_candidate_urls_start_from_the_host_that_resolved():
    assert candidate_urls("shop.com", "shop.com")[0] == "https://shop.com/"
    assert candidate_urls("shop.com", "www.shop.com")[0] == "https://www.shop.com/"
    assert candidate_urls("shop.com", None) == [
        "https://shop.com/", "http://shop.com/", "https://www.shop.com/"]


# --- stage 4: scoring and tiering -----------------------------------------

def make_row(**overrides):
    row = {"status": config.STATUS_LIVE, "platform": "Shopify", "has_cart": 1,
           "has_product_schema": 1, "created_year": 2024, "has_mx": 1,
           "response_time_ms": 500, "is_role": 0, "is_freemail": 0,
           "is_parked": 0, "http_status": 200, "jurisdiction": "PERMISSIVE"}
    row.update(overrides)
    return row


def test_score_is_bounded_and_ordered():
    assert score_row(make_row()) == 100
    weak = make_row(platform=None, has_product_schema=0, created_year=2016,
                    has_mx=0, response_time_ms=9000, is_role=1)
    assert 0 <= score_row(weak) < score_row(make_row())


def test_platform_fit_dominates():
    shopify = score_row(make_row(platform="Shopify"))
    wix = score_row(make_row(platform="Wix"))
    none = score_row(make_row(platform=None))
    assert shopify > wix > none


def test_role_inbox_is_a_small_penalty():
    assert score_row(make_row()) - score_row(make_row(is_role=1)) == 5


def test_recency_points_are_monotonic():
    assert (recency_points(2025) >= recency_points(2023)
            >= recency_points(2021) >= recency_points(2018) == 0)
    assert recency_points(None) == 0


def test_permissive_store_is_cleared_for_email():
    tier, score, _ = assign(make_row(), 55)
    assert tier == TIER1 and score >= 55


def test_uk_is_emailable_but_eu_and_canada_are_ads_only():
    assert assign(make_row(jurisdiction="PECR"), 55)[0] == TIER1
    for jurisdiction in ("EU", "CASL", "OTHER"):
        tier, _, reason = assign(make_row(jurisdiction=jurisdiction), 55)
        assert tier == TIER2, jurisdiction
        assert "never email" in reason or "no cold email" in reason


def test_freemail_never_reaches_the_email_tier():
    tier, _, reason = assign(make_row(status=config.STATUS_FREEMAIL), 55)
    assert tier == TIER2
    assert "never email" in reason


def test_freemail_on_a_permissive_tld_is_still_ads_only():
    """gmail.com is a .com, which must not make its contacts emailable."""
    row = make_row(is_freemail=1, jurisdiction="PERMISSIVE")
    assert assign(row, 55)[0] == TIER2


def test_no_mx_cannot_be_emailed():
    assert assign(make_row(has_mx=0), 55)[0] == TIER2


def test_low_score_goes_to_nurture():
    weak = make_row(platform=None, has_product_schema=0, created_year=2016,
                    response_time_ms=8000)
    tier, score, _ = assign(weak, 55)
    assert tier == TIER3 and score < 55


@pytest.mark.parametrize("row, fragment", [
    (make_row(is_parked=1), "parked"),
    (make_row(status=config.STATUS_PARKED), "parked"),
    (make_row(has_cart=0), "cart"),
    (make_row(http_status=404), "non-200"),
    (make_row(status=config.STATUS_DEAD), "resolve"),
    (make_row(status=config.STATUS_MX_ONLY), "no website"),
    (make_row(status=config.STATUS_BLOCKED), "robots"),
])
def test_disqualifying_conditions_are_dropped_with_a_reason(row, fragment):
    tier, _, reason = assign(row, 55)
    assert tier == DROPPED
    assert fragment in reason


def test_uncrawled_targets_get_no_tier():
    """An unfinished crawl must never read as a rejection."""
    tier, score, reason = assign(make_row(status=config.STATUS_PENDING), 55)
    assert tier is None and score is None and reason is None


def test_aged_out_targets_are_nurtured_not_dropped():
    tier, _, reason = assign(make_row(status=config.STATUS_AGED_OUT), 55)
    assert tier == TIER3
    assert "not fetched" in reason


# --- schema ----------------------------------------------------------------

def test_missing_columns_are_added_to_an_existing_database(tmp_path):
    """A half-enriched database represents hours of lookups; a schema change
    must not force the user to start over."""
    path = str(tmp_path / "old.db")
    legacy = sqlite3.connect(path)
    legacy.execute("CREATE TABLE targets (target_key TEXT PRIMARY KEY, "
                   "domain TEXT NOT NULL, email TEXT NOT NULL)")
    legacy.execute("INSERT INTO targets VALUES ('s.com', 's.com', 'a@s.com')")
    legacy.commit()
    legacy.close()

    conn = db.connect(path)
    columns = {r["name"] for r in conn.execute("PRAGMA table_info(targets)")}
    assert {"a_host", "platform", "tier", "tier_reason"} <= columns
    assert conn.execute("SELECT COUNT(*) FROM targets").fetchone()[0] == 1
    conn.close()


def test_tier3_never_contains_a_target_that_cannot_be_emailed():
    """Tier 3 is a holding pen for future email, so nothing that may never be
    emailed is allowed to sit in it."""
    weak = dict(platform=None, has_product_schema=0, created_year=2016,
                response_time_ms=8000)
    assert assign(make_row(**weak), 55)[0] == TIER3          # permissive: fine
    assert assign(make_row(jurisdiction="EU", **weak), 55)[0] == TIER2
    assert assign(make_row(jurisdiction="CASL", **weak), 55)[0] == TIER2
    assert assign(make_row(is_freemail=1, **weak), 55)[0] == TIER2
    assert assign(make_row(has_mx=0, **weak), 55)[0] == TIER2


def test_an_aged_out_eu_contact_is_ads_only_not_nurture():
    tier, _, reason = assign(
        make_row(status=config.STATUS_AGED_OUT, jurisdiction="EU"), 55)
    assert tier == TIER2
    assert "no cold email" in reason


def test_uncrawled_targets_are_not_given_a_score():
    """A score built from absent signals would look like a judgement."""
    for status in (config.STATUS_AGED_OUT, config.STATUS_FREEMAIL):
        assert assign(make_row(status=status), 55)[1] == 0


# --- stage 2: a timeout is not an answer -----------------------------------

class FakeAnswer(list):
    pass


class FakeRecord:
    def __init__(self, exchange, preference):
        self.exchange, self.preference = exchange, preference


class FakeResolver:
    """Resolver whose answers are scripted per (name, rdtype)."""

    def __init__(self, script):
        self.script = script

    async def resolve(self, name, rdtype):
        outcome = self.script.get((name, rdtype), dns.resolver.NXDOMAIN())
        if isinstance(outcome, Exception):
            raise outcome
        return outcome


async def check(script, domain="store.com"):
    return await check_domain(FakeResolver(script), domain, retries=0)


def test_dns_timeout_is_not_recorded_as_a_dead_domain():
    """The failure mode that deleted 26k real stores from a run: a saturated
    resolver times out, and every timeout gets written down as 'dead'."""
    result = asyncio.run(check({
        ("store.com", "A"): dns.exception.Timeout(),
        ("store.com", "MX"): dns.exception.Timeout(),
    }))
    assert result["inconclusive"] is True


def test_an_authoritative_nxdomain_is_conclusive():
    result = asyncio.run(check({}))  # everything NXDOMAINs
    assert result["inconclusive"] is False
    assert result["status"] == config.STATUS_DEAD
    assert result["dns_resolves"] == 0


def test_a_resolving_domain_is_conclusive_even_if_mx_timed_out():
    result = asyncio.run(check({
        ("store.com", "A"): FakeAnswer(["1.2.3.4"]),
        ("store.com", "MX"): dns.exception.Timeout(),
    }))
    assert result["inconclusive"] is False
    assert result["dns_resolves"] == 1


def test_mx_only_domain_has_no_website():
    result = asyncio.run(check({
        ("store.com", "MX"): FakeAnswer([FakeRecord("mx.mail.com.", 10)]),
    }))
    assert result["status"] == config.STATUS_MX_ONLY
    assert result["mx_provider"] == "mx.mail.com"


def test_lowest_preference_mx_wins():
    result = asyncio.run(check({
        ("store.com", "MX"): FakeAnswer([FakeRecord("backup.mail.com.", 50),
                                         FakeRecord("primary.mail.com.", 10)]),
    }))
    assert result["mx_provider"] == "primary.mail.com"


def test_www_fallback_rescues_a_www_only_store():
    result = asyncio.run(check({
        ("www.store.com", "A"): FakeAnswer(["1.2.3.4"]),
    }))
    assert result["dns_resolves"] == 1
    assert result["a_host"] == "www.store.com"


# --- stage 3: a timeout is not an answer either -----------------------------

def scripted_fetch(results):
    """Replacement for stage3.fetch that returns a canned result per URL."""
    async def _fetch(session, url, retry_timeout=True, timeout=None):
        for needle, result in results:
            if needle in url:
                return dict(result)
        return {"status": None, "final_url": None, "headers": {}, "body": b"",
                "content_length": None, "truncated": False, "elapsed_ms": 1,
                "error": "timeout"}
    return _fetch


def ok(body=b"<html></html>", status=200, headers=None):
    return {"status": status, "final_url": "https://store.com/", "headers": headers or {},
            "body": body, "content_length": len(body), "truncated": False,
            "elapsed_ms": 120, "error": None}


def refused():
    return {"status": None, "final_url": None, "headers": {}, "body": b"",
            "content_length": None, "truncated": False, "elapsed_ms": 5,
            "error": "ClientConnectorError"}


def run_check(monkeypatch, results, a_host="store.com"):
    import pipeline.stage3_http as stage3
    monkeypatch.setattr(stage3, "fetch", scripted_fetch(results))
    row = {"domain": "store.com", "a_host": a_host}
    return asyncio.run(stage3.check_domain(object(), row, ignore_robots=True))


def test_a_fetch_timeout_is_not_recorded_as_a_dead_store(monkeypatch):
    record = run_check(monkeypatch, [])  # everything times out
    assert record["inconclusive"] is True


def test_a_refused_connection_is_conclusively_dead(monkeypatch):
    record = run_check(monkeypatch, [("store.com", refused())])
    assert record["inconclusive"] is False
    assert record["status"] == config.STATUS_DEAD


def test_a_live_store_is_fingerprinted(monkeypatch):
    body = (b'<html lang="en"><title>My Shop</title>'
            b'<script src="https://cdn.shopify.com/s/x.js"></script>'
            b'<a href="/cart">Cart</a>'
            b'<script type="application/ld+json">{"@type":"Product"}</script></html>')
    record = run_check(monkeypatch, [("store.com", ok(body))])
    assert record["status"] == config.STATUS_LIVE
    assert record["platform"] == "Shopify"
    assert record["has_cart"] == 1 and record["has_product_schema"] == 1
    assert record["page_title"] == "My Shop" and record["lang"] == "en"
    assert record["inconclusive"] is False


def test_a_parked_page_is_not_live(monkeypatch):
    record = run_check(monkeypatch,
                       [("store.com", ok(b"<title>This domain is for sale</title>"))])
    assert record["status"] == config.STATUS_PARKED


def test_a_non_200_response_is_an_error_not_a_live_store(monkeypatch):
    record = run_check(monkeypatch, [("store.com", ok(status=503))])
    assert record["status"] == config.STATUS_ERROR
    assert record["http_status"] == 503


def test_http_fallback_is_used_when_https_is_refused(monkeypatch):
    record = run_check(monkeypatch, [("https://store.com", refused()),
                                     ("http://store.com", ok())])
    assert record["status"] == config.STATUS_LIVE


def test_an_nxdomain_a_record_settles_the_domain_even_if_mx_flaked():
    """The domain does not exist, so it has no mail either -- looking again
    three times over would just cost queries."""
    result = asyncio.run(check({
        ("store.com", "MX"): dns.exception.Timeout(),
    }))
    assert result["inconclusive"] is False
    assert result["status"] == config.STATUS_DEAD


def test_a_timed_out_a_record_is_never_settled():
    result = asyncio.run(check({
        ("store.com", "A"): dns.exception.Timeout(),
        ("store.com", "MX"): dns.resolver.NoAnswer(),
    }))
    assert result["inconclusive"] is True


# --- stage 3: robots.txt ---------------------------------------------------

def robots(body=b"", status=200):
    return {"status": status, "final_url": None, "headers": {}, "body": body,
            "content_length": len(body), "truncated": False, "elapsed_ms": 3,
            "error": None}


def run_with_robots(monkeypatch, results, a_host="store.com"):
    import pipeline.stage3_http as stage3
    monkeypatch.setattr(stage3, "fetch", scripted_fetch(results))
    row = {"domain": "store.com", "a_host": a_host}
    return asyncio.run(stage3.check_domain(object(), row, ignore_robots=False))


def test_a_disallowed_root_is_never_fetched(monkeypatch):
    record = run_with_robots(monkeypatch, [
        ("robots.txt", robots(b"User-agent: *\nDisallow: /")),
        ("store.com", ok()),
    ])
    assert record["status"] == config.STATUS_BLOCKED
    assert record["robots_allowed"] == 0
    assert record["http_status"] is None  # the page was not requested


def test_an_allowed_root_is_fetched(monkeypatch):
    record = run_with_robots(monkeypatch, [
        ("robots.txt", robots(b"User-agent: *\nDisallow: /admin\nAllow: /")),
        ("store.com", ok()),
    ])
    assert record["status"] == config.STATUS_LIVE
    assert record["robots_allowed"] == 1


def test_a_404_on_robots_means_no_rules(monkeypatch):
    record = run_with_robots(monkeypatch, [
        ("robots.txt", robots(b"Not Found", status=404)),
        ("store.com", ok()),
    ])
    assert record["status"] == config.STATUS_LIVE


def test_a_5xx_on_robots_is_treated_as_a_refusal(monkeypatch):
    record = run_with_robots(monkeypatch, [
        ("robots.txt", robots(b"", status=503)),
        ("store.com", ok()),
    ])
    assert record["status"] == config.STATUS_BLOCKED


def test_robots_is_consulted_on_the_origin_actually_fetched(monkeypatch):
    """An http-only store's rules must be read too, not skipped because the
    https origin never answered."""
    record = run_with_robots(monkeypatch, [
        ("https://store.com/robots.txt", refused()),
        ("https://store.com/", refused()),
        ("http://store.com/robots.txt", robots(b"User-agent: *\nDisallow: /")),
        ("http://store.com/", ok()),
    ])
    assert record["status"] == config.STATUS_BLOCKED


# --- stage 1: the input shapes, end to end ---------------------------------

def build_workbook(path):
    """A workbook mixing both observed shapes, plus the messes seen in the data."""
    import openpyxl as opx
    wb = opx.Workbook()

    headed = wb.active
    headed.title = "Sheet1"
    headed.append(["created", "emails"])
    headed.append([dt.datetime(2024, 6, 7), "info@alpha.com"])
    # Several addresses packed into one cell, colon separated.
    headed.append([dt.datetime(2023, 1, 2), "a@beta.com:info@beta.com:sales@beta.com"])
    # An older contact at a domain that also appears newer, further down.
    headed.append([dt.datetime(2017, 3, 4), "old@alpha.com"])
    headed.append([dt.datetime(2025, 2, 1), "hola@gamma.es"])
    headed.append([dt.datetime(2016, 9, 23)])          # date, no email
    headed.append([dt.datetime(2020, 1, 1), "not an email at all"])

    ragged = wb.create_sheet("Sheet3")
    ragged.append(["info@delta.co.uk", "support@delta.co.uk"])   # 2 columns
    ragged.append(["shop@epsilon.ca"])                            # 1 column
    ragged.append([None, "mailto:kontakt@zeta.de", None, "x"])    # gaps and junk
    ragged.append(["one@eta.com"] + [None] * 12 + ["two@theta.com"])  # 14 wide
    ragged.append(["mailto"])                                     # bare junk
    wb.save(path)


def test_both_input_shapes_are_ingested(tmp_path):
    from pipeline.stage1_ingest import main as stage1
    data = tmp_path / "data"
    data.mkdir()
    build_workbook(str(data / "part1_of_4.xlsx"))
    db_path = str(tmp_path / "t.db")

    assert stage1(["--db", db_path, "--data-dir", str(data)]) == 0

    conn = db.connect(db_path)
    rows = {r["domain"]: r for r in conn.execute("SELECT * FROM targets")}

    # Every domain from both sheets, and nothing invented from the junk cells.
    assert set(rows) == {"alpha.com", "beta.com", "gamma.es", "delta.co.uk",
                         "epsilon.ca", "zeta.de", "eta.com", "theta.com"}

    # Domain dedupe kept the newest contact of the three at beta.com, and
    # preferred the 2024 alpha.com contact over the 2017 one.
    assert rows["beta.com"]["email_count"] == 3
    assert rows["alpha.com"]["created_year"] == 2024
    assert rows["alpha.com"]["email"] == "info@alpha.com"

    # The headerless sheet carries no dates at all.
    assert rows["eta.com"]["created"] is None

    # Classification survived the trip through the workbook.
    assert rows["delta.co.uk"]["jurisdiction"] == "PECR"
    assert rows["zeta.de"]["jurisdiction"] == "EU"
    assert rows["epsilon.ca"]["jurisdiction"] == "CASL"
    assert rows["gamma.es"]["is_role"] == 1
    conn.close()


def test_a_second_ingest_is_a_no_op(tmp_path):
    from pipeline.stage1_ingest import main as stage1
    data = tmp_path / "data"
    data.mkdir()
    build_workbook(str(data / "part1_of_4.xlsx"))
    db_path = str(tmp_path / "t.db")

    stage1(["--db", db_path, "--data-dir", str(data)])
    conn = db.connect(db_path)
    before = conn.execute("SELECT SUM(email_count) FROM targets").fetchone()[0]

    stage1(["--db", db_path, "--data-dir", str(data)])
    after = conn.execute("SELECT SUM(email_count) FROM targets").fetchone()[0]
    assert after == before, "re-running stage 1 must not double-count contacts"
    conn.close()


# --- stage 4: the workbooks themselves -------------------------------------

def seed_scored_targets(conn):
    """Four targets, one destined for each workbook."""
    rows = [
        # a live Shopify store on a .com: tier 1
        ("good.com", "good.com", "team@good.com", "PERMISSIVE", "com", 0, 0, 2024,
         1, 1, "live", "Shopify", 1, 1, 0, 200, 300),
        # the same store, but German: tier 2
        ("gute.de", "gute.de", "info@gute.de", "EU", "de", 0, 1, 2024,
         1, 1, "live", "Shopify", 1, 1, 0, 200, 300),
        # live but nothing going for it: tier 3
        ("meh.com", "meh.com", "info@meh.com", "PERMISSIVE", "com", 0, 1, 2016,
         1, 1, "live", None, 1, 0, 0, 200, 8000),
        # parked: dropped
        ("gone.com", "gone.com", "info@gone.com", "PERMISSIVE", "com", 0, 1, 2024,
         1, 1, "parked", None, 0, 0, 1, 200, 300),
    ]
    conn.executemany(
        "INSERT INTO targets (target_key, domain, email, jurisdiction, tld, "
        "is_freemail, is_role, created_year, dns_resolves, has_mx, status, "
        "platform, has_cart, has_product_schema, is_parked, http_status, "
        "response_time_ms) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", rows)
    conn.commit()


def test_stage4_writes_one_workbook_per_tier(tmp_path):
    import openpyxl as opx
    from pipeline.stage4_score import main as stage4
    db_path = str(tmp_path / "t.db")
    conn = db.connect(db_path)
    seed_scored_targets(conn)
    conn.close()

    out = tmp_path / "out"
    assert stage4(["--db", db_path, "--out-dir", str(out)]) == 0

    expected = {"tier1_email.xlsx": "good.com", "tier2_ads.xlsx": "gute.de",
                "tier3_nurture.xlsx": "meh.com", "dropped.xlsx": "gone.com"}
    for filename, domain in expected.items():
        book = opx.load_workbook(str(out / filename))
        sheet = book.worksheets[0]
        header = [c.value for c in sheet[3]]
        assert header[:10] == ["domain", "email", "platform", "score",
                               "created_year", "jurisdiction", "http_status",
                               "has_cart", "page_title", "final_url"], filename
        body = [r[0] for r in sheet.iter_rows(min_row=4, values_only=True)]
        assert body == [domain], filename
        book.close()

    # dropped.xlsx carries the reason; the others do not.
    dropped = opx.load_workbook(str(out / "dropped.xlsx"))
    sheet = dropped.worksheets[0]
    assert [c.value for c in sheet[3]][-1] == "reason"
    assert "parked" in sheet.cell(row=4, column=11).value
    dropped.close()


def test_workbooks_are_sorted_by_descending_score(tmp_path):
    import openpyxl as opx
    from pipeline.stage4_score import main as stage4
    db_path = str(tmp_path / "t.db")
    conn = db.connect(db_path)
    conn.executemany(
        "INSERT INTO targets (target_key, domain, email, jurisdiction, tld, "
        "is_freemail, is_role, created_year, dns_resolves, has_mx, status, "
        "platform, has_cart, has_product_schema, is_parked, http_status, "
        "response_time_ms) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [(f"s{i}.com", f"s{i}.com", f"team@s{i}.com", "PERMISSIVE", "com", 0, 0,
          year, 1, 1, "live", platform, 1, 1, 0, 200, 300)
         for i, (year, platform) in enumerate(
             [(2020, "Wix"), (2024, "Shopify"), (2022, "Magento")])])
    conn.commit()
    conn.close()

    out = tmp_path / "out"
    stage4(["--db", db_path, "--out-dir", str(out)])
    book = opx.load_workbook(str(out / "tier1_email.xlsx"))
    scores = [r[3] for r in book.worksheets[0].iter_rows(min_row=4, values_only=True)]
    assert scores == sorted(scores, reverse=True)
    book.close()


def test_a_limited_ingest_is_not_recorded_as_complete(tmp_path):
    """Otherwise a --limit smoke test makes the real run skip the file."""
    from pipeline.stage1_ingest import main as stage1
    data = tmp_path / "data"
    data.mkdir()
    build_workbook(str(data / "part1_of_4.xlsx"))
    db_path = str(tmp_path / "t.db")

    stage1(["--db", db_path, "--data-dir", str(data), "--limit", "2"])
    conn = db.connect(db_path)
    assert conn.execute("SELECT COUNT(*) FROM ingest_log").fetchone()[0] == 0
    partial = conn.execute("SELECT COUNT(*) FROM targets").fetchone()[0]
    assert partial < 8

    stage1(["--db", db_path, "--data-dir", str(data)])
    assert conn.execute("SELECT COUNT(*) FROM targets").fetchone()[0] == 8
    assert conn.execute("SELECT COUNT(*) FROM ingest_log").fetchone()[0] == 1
    conn.close()


def test_force_clears_the_attempt_counters(tmp_path):
    """--force asks for a fresh verdict, so a domain that timed out twice
    before must not be settled as dead by its first timeout of this run."""
    import types
    from pipeline.stage2_dns import pending_domains
    from pipeline.stage3_http import pending_rows

    db_path = str(tmp_path / "t.db")
    conn = db.connect(db_path)
    conn.execute("INSERT INTO targets (target_key, domain, email, dns_attempts, "
                 "http_attempts, dns_resolves, created_year, dns_checked_at, "
                 "http_checked_at) VALUES ('s.com','s.com','a@s.com',2,2,1,2024,"
                 "'2026-01-01','2026-01-01')")
    conn.commit()

    pending_domains(conn, force=True, limit=None)
    assert conn.execute("SELECT dns_attempts FROM targets").fetchone()[0] == 0

    args = types.SimpleNamespace(force=True, limit=None, min_year=2020)
    pending_rows(conn, args)
    assert conn.execute("SELECT http_attempts FROM targets").fetchone()[0] == 0
    conn.close()


def test_resume_leaves_the_attempt_counters_alone(tmp_path):
    import types
    from pipeline.stage2_dns import pending_domains

    db_path = str(tmp_path / "t.db")
    conn = db.connect(db_path)
    conn.execute("INSERT INTO targets (target_key, domain, email, dns_attempts) "
                 "VALUES ('s.com','s.com','a@s.com',2)")
    conn.commit()
    pending_domains(conn, force=False, limit=None)
    assert conn.execute("SELECT dns_attempts FROM targets").fetchone()[0] == 2
    conn.close()
